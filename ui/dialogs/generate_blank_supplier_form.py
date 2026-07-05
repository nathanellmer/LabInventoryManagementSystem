from openpyxl import load_workbook
from datetime import datetime
from PySide6.QtWidgets import QDialog, QHBoxLayout, QVBoxLayout, QListWidget
from PySide6.QtCore import Qt
from ui.custom_widgets.general.header_widgets import FormHeaderWidget
from ui.custom_widgets.general.footer_widgets import FormFooterWidget
from ui.custom_widgets.general.button_widgets import MainMenuButton, MainMenuLinkButton
from ui.custom_widgets.general.form_widgets import FormLabelComboWidgetWide, FormSearchPanelWidget, FormLabelTextWidget
from ui.dialogs.add_grant_codes_dialog import AddGrantCodesDialog
from ui.dialogs.choice_selection_dialog import ChoiceSelectionDialogTwo, ChoiceSelectionDialogFour
from ui.dialogs.show_item_dialog import ShowItemDialog
from core.control_functions import controller, db_info
from core.db_get_functions import get_all_suppliers,  get_grant_code_info_by_grant_code_name, get_grant_code_info_by_grant_code_owner, get_user_info_by_username, get_grant_code_info_by_id, get_supplier_info_by_name
from core.utility_functions import resource_path


class GenerateBlankSupplierFormDialog(QDialog):
    def __init__(self):
        super().__init__()

        # Set the window size and title
        self.resize(500, 400)
        self.setWindowTitle("Generate Blank Supplier Forms")

        # Initialise the GenerateBlankSupplierFormDialog layout and its margins (left, top, right, bottom)
        gen_sup_forms_dialog_layout = QVBoxLayout()
        gen_sup_forms_dialog_layout.setContentsMargins(20, 20, 20, 20)

        # Add the header widget to the generate order forms dialog layout
        gen_sup_forms_dialog_layout.addWidget(FormHeaderWidget("Generate Blank Supplier Forms:"))

        # Add stretch
        gen_sup_forms_dialog_layout.addStretch()

        # Combo box for selecting the supplier to generate the order form for
        item_search_layout = QHBoxLayout()
        item_search_layout.addStretch()
        self.item_supplier = FormLabelComboWidgetWide("Supplier:")
        self.load_suppliers()
        item_search_layout.addStretch()
        gen_sup_forms_dialog_layout.addLayout(self.item_supplier)

        # Add stretch
        gen_sup_forms_dialog_layout.addStretch()

        # Add a search panel for the grant codes
        grant_code_search_layout = QHBoxLayout()
        grant_code_search_layout.addStretch()
        self.grant_code_search_panel = FormSearchPanelWidget("Search for a grant code:", "Grant code or owner:", self.btn_grant_code_search)
        grant_code_search_layout.addWidget(self.grant_code_search_panel)
        grant_code_search_layout.addStretch()
        gen_sup_forms_dialog_layout.addLayout(grant_code_search_layout)

        # Add a link button widget for missing grant code information
        self.btn_missing_grant_code = MainMenuLinkButton("I cannot find the grant code...", self.open_add_grant_code_dialog)
        gen_sup_forms_dialog_layout.addWidget(self.btn_missing_grant_code)

        # Add spacing
        gen_sup_forms_dialog_layout.addSpacing(10)

        # Add button for order form generation
        gen_sup_forms_dialog_layout.addWidget(MainMenuButton("Generate Order Form", self.btn_generate_order_form), alignment=Qt.AlignCenter)

        # Add stretch
        gen_sup_forms_dialog_layout.addStretch()

        # Add the universal footer widget to the generate order forms dialog layout
        controller.close_all_windows.connect(self.close)
        gen_sup_forms_dialog_layout.addWidget(FormFooterWidget())

        # Set the main layout for the window
        self.setLayout(gen_sup_forms_dialog_layout)


    def load_suppliers(self):
        # Clear suppliers from the dropdown
        self.item_supplier.cmb.clear()

        # Get the list of suppliers from the database and add them to the dropdown
        suppliers = get_all_suppliers()
        suppliers.insert(0, "Select Supplier...")
        self.item_supplier.cmb.addItems(suppliers)


    def btn_grant_code_search(self):
        # Gather the data from the form
        field_values = []

        for field in self.grant_code_search_panel.findChildren(FormLabelTextWidget):
            field_values.append(field.txt.text())

        # Search the database for the grant code by code
        grant_codes_by_name = get_grant_code_info_by_grant_code_name(field_values[0])

        # Search the database for the grant code by owner
        grant_codes_by_owner = get_grant_code_info_by_grant_code_owner(field_values[0])

        choice_selection_dialog = ChoiceSelectionDialogTwo("Multiple Grant Codes Found", "Multiple grant codes were found matching your search. Please select one:", "Grant Codes by Name:", "Grant Codes by Owner:", [grant_code[1] for grant_code in grant_codes_by_name], [grant_code[1] for grant_code in grant_codes_by_owner])
        if choice_selection_dialog.exec() == QDialog.Accepted:
            selected_optn = choice_selection_dialog.selected_option
            selected_idx = choice_selection_dialog.selected_idx

            if selected_optn == 1:
                selected_grant_code = grant_codes_by_name[selected_idx]
            else:
                selected_grant_code = grant_codes_by_owner[selected_idx]

            self.grant_code_id = selected_grant_code[0] 
            
            for idx, field in enumerate(self.grant_code_search_panel.findChildren(FormLabelTextWidget)):
                field.txt.setText(selected_grant_code[idx + 1])


    def open_add_grant_code_dialog(self):
        add_grant_code_dialog = AddGrantCodesDialog()
        add_grant_code_dialog.exec()


    def btn_generate_order_form(self):
        supplier_name = self.item_supplier.cmb.currentText()
        supplier_info = get_supplier_info_by_name(supplier_name)

        self.populate_order_form(supplier_info)


    def populate_order_form(self, supplier_info):
        # Gather all the information for the form
        grant_code_info = get_grant_code_info_by_id(self.grant_code_id)
        user_info = get_user_info_by_username(controller.logged_in_user)

        wb = load_workbook(resource_path("templates/Engineering_Order_Form_Template.xlsx"))
        ws = wb.active

        # Populate the supplier details
        ws[f"B10"] = supplier_info[1]
        ws[f"B11"] = supplier_info[3]
        ws[f"B12"] = supplier_info[4]
        ws[f"B13"] = supplier_info[5]
        ws[f"B14"] = supplier_info[6]
        ws[f"B16"] = supplier_info[7]
        ws[f"B17"] = supplier_info[8]
        ws[f"B18"] = supplier_info[9]
        ws[f"A50"] = supplier_info[2]
        ws[f"A53"] = "This is a trusted supplier."

        # Populate delivery details
        ws[f"F10"] = user_info[1]
        ws[f"F11"] = user_info[2]
        ws[f"F13"] = user_info[3]
        ws[f"F14"] = "Engineering Service Yard"
        ws[f"F15"] = "Swansea University Bay Campus"
        ws[f"F16"] = "Skewen"
        ws[f"F17"] = "Swansea"
        ws[f"F18"] = "SA1 8EN"

        # Populate the grant code details
        ws[f"A56"] = grant_code_info[1]
        ws[f"H55"] = datetime.now().strftime("%d/%m/%Y")
        ws[f"H56"] = user_info[1]
        ws[f"H57"] = grant_code_info[2]

        wb.save(f"{db_info.SAVE_PATH}Blank_Engineering_Order_Form_{supplier_info[1]}.xlsx")
