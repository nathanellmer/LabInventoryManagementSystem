from openpyxl import load_workbook
from datetime import datetime
from PySide6.QtWidgets import QDialog, QHBoxLayout, QVBoxLayout, QScrollArea, QWidget
from PySide6.QtCore import Qt
from ui.custom_widgets.general.header_widgets import FormHeaderWidget
from ui.custom_widgets.general.footer_widgets import FormFooterWidget
from ui.custom_widgets.general.button_widgets import MainMenuButton, MainMenuLinkButton
from ui.custom_widgets.general.form_widgets import FormLabelCheckBox, FormLabelComboWidgetWide, FormLabelHyperlinkButton, FormLabelLinkButton, FormLabelTextWidgetExtraWide, FormLabelTextWidgetWide, FormSearchPanelWidget, FormLabelTextWidget
from ui.custom_widgets.window_dialog_panels.db_item_widgets import ItemsChemicalDisplayPanelWidget, PictogramWidget, ReadOnlyItemsPanelWidget
from ui.dialogs.choice_selection_dialog import ChoiceSelectionDialogTwo, ChoiceSelectionDialogFour
from ui.dialogs.add_grant_codes_dialog import AddGrantCodesDialog
from core.control_functions import controller, db_info
from core.db_get_functions import get_item_info_by_product_code, get_item_info_by_supplier, get_item_info_by_description, get_item_info_by_name, get_supplier_info_by_id, get_storage_location_info_by_id, get_grant_code_info_by_grant_code_name, get_grant_code_info_by_grant_code_owner, get_item_info_by_id, get_user_info_by_username, get_grant_code_info_by_id
from core.utility_functions import resource_path

class SearchItemDialog(QDialog):
    def __init__(self):
        super().__init__()

        # Set the window size and title
        self.resize(1250, 600)
        self.setWindowTitle("Search Items")
        self.setMaximumHeight(600)

        # Initialise the SearchItemDialog layout and its margins (left, top, right, bottom)
        search_item_dialog_layout = QVBoxLayout()
        search_item_dialog_layout.setContentsMargins(20, 20, 20, 20)

        # Add the header widget to the search item dialog layout
        search_item_dialog_layout.addWidget(FormHeaderWidget("Search Item Information:"))

        # Add stretch
        search_item_dialog_layout.addStretch()

        # Add a search panel
        item_search_layout = QHBoxLayout()
        item_search_layout.addStretch()
        self.item_search_panel = FormSearchPanelWidget("Search for an item:", "Name or other information:", self.btn_item_search)
        item_search_layout.addWidget(self.item_search_panel)
        item_search_layout.addStretch()
        search_item_dialog_layout.addLayout(item_search_layout)

        # Add stretch
        search_item_dialog_layout.addStretch()

        # Add the items panel widget to the search item dialog layout
        self.item_form = ReadOnlyItemsPanelWidget()
        search_item_dialog_layout.addWidget(self.item_form, alignment=Qt.AlignCenter)

        # Add spacing
        search_item_dialog_layout.addSpacing(10)

        # Add the chemical section widget to the search item dialog layout
        self.chemical_flag_1 = False
        self.chemical_flag_2 = False
        self.chemical_section = ItemsChemicalDisplayPanelWidget("", "")
        search_item_dialog_layout.addWidget(self.chemical_section)
        self.chemical_section.hide()

        # Add spacing
        search_item_dialog_layout.addSpacing(10)

        # Add a search panel for the grant codes
        grant_code_search_layout = QHBoxLayout()
        grant_code_search_layout.addStretch()
        self.grant_code_search_panel = FormSearchPanelWidget("Search for a grant code:", "Grant code or owner:", self.btn_grant_code_search)
        grant_code_search_layout.addWidget(self.grant_code_search_panel)
        grant_code_search_layout.addStretch()
        search_item_dialog_layout.addLayout(grant_code_search_layout)

        # Add a link button widget for missing grant code information
        self.btn_missing_grant_code = MainMenuLinkButton("I cannot find the grant code...", self.open_add_grant_code_dialog)
        search_item_dialog_layout.addWidget(self.btn_missing_grant_code)

        # Add spacing
        search_item_dialog_layout.addSpacing(20)

        # Add a proceed button
        search_item_dialog_layout.addWidget(MainMenuButton("Export Form", self.btn_proceed), alignment=Qt.AlignCenter)

        # Add stretch
        search_item_dialog_layout.addStretch()

        # Add the universal footer widget to the search item dialog layout
        controller.close_all_windows.connect(self.close)
        search_item_dialog_layout.addWidget(FormFooterWidget())

        # Create a container widget for the scroll area and set its layout
        container = QWidget()
        container.setLayout(search_item_dialog_layout)

        # Create a scroll area and set the container widget as its widget
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(container)

        # Set the main layout for the window
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll_area)


    def btn_item_search(self):
        # Gather the data from the form
        field_values = []

        for field in self.item_search_panel.findChildren(FormLabelTextWidget):
            field_values.append(field.txt.text())

        # Search the database for the item by name
        items_by_name = get_item_info_by_name(field_values[0])

        # Search the database for the item by supplier
        items_by_supplier = get_item_info_by_supplier(field_values[0])

        # Search the database for the item by product code
        items_by_ref = get_item_info_by_product_code(field_values[0])

        # Search the database for the item by description
        items_by_description = get_item_info_by_description(field_values[0])

        choice_selection_dialog = ChoiceSelectionDialogFour("Multiple Items Found", "Multiple items were found matching your search. Please select one:", "Items by Name:", "Items by Supplier:", "Items by Product Code:", "Items by Description:", [item[1] for item in items_by_name], [item[1] for item in items_by_supplier], [item[1] for item in items_by_ref], [item[1] for item in items_by_description])
        if choice_selection_dialog.exec() == QDialog.Accepted:
            selected_optn = choice_selection_dialog.selected_option
            selected_idx = choice_selection_dialog.selected_idx

            if selected_optn == 1:
                selected_item = items_by_name[selected_idx]
            elif selected_optn == 2:
                selected_item = items_by_supplier[selected_idx]
            elif selected_optn == 3:
                selected_item = items_by_ref[selected_idx]
            elif selected_optn == 4:
                selected_item = items_by_description[selected_idx]

            self.item_id = selected_item[0]

            info_idx = [1, 2, 3, 4, 5, 6, 8, 7, 12, 13]
            for idx, field in enumerate(self.item_form.findChildren(FormLabelTextWidgetWide)):
                if idx == 1:
                    info = get_supplier_info_by_id(selected_item[info_idx[idx]])
                    field.txt.setText(info[1])
                elif idx == 7:
                    info = get_storage_location_info_by_id(selected_item[info_idx[idx]])
                    field.txt.setText(info[1])
                elif idx == 4:
                    field.txt.setText(str(selected_item[info_idx[idx]]))
                elif idx == 5:
                    field.txt.setText(str(selected_item[info_idx[idx]]))
                elif idx == 6:
                    field.txt.setText(str(selected_item[info_idx[idx]]))
                else:
                    field.txt.setText(selected_item[info_idx[idx]])

            info_idx = 9
            field = self.item_form.findChildren(FormLabelHyperlinkButton)[0]
            field.btn.setText(selected_item[info_idx])
            field.set_url(selected_item[info_idx])

            info_idx = 10
            field = self.item_form.findChildren(FormLabelCheckBox)[0]
            field.chb.setChecked(selected_item[info_idx])
            field.chb.setEnabled(False)

            self.originator_id = selected_item[11]

            # If the item is a chemical, show the chemical section and populate the chemical information
            if selected_item[12] == "Hazardous Chemical":
                self.chemical_section.show()
                self.chemical_flag_1 = True

                info_idx = [14]
                for idx, field in enumerate(self.chemical_section.findChildren(FormLabelTextWidgetExtraWide)):
                    field.txt.setText(selected_item[info_idx[idx]])
                    field.txt.setReadOnly(True)

                info_idx = [15, 16]
                for idx, field in enumerate(self.chemical_section.findChildren(FormLabelLinkButton)):
                    field.btn.setText(selected_item[info_idx[idx]])

                info_idx = [17, 18, 19, 20, 21, 22, 23, 24, 25]
                for idx, field in enumerate(self.chemical_section.findChildren(PictogramWidget)):
                    field.chb.setChecked(selected_item[info_idx[idx]])
                    field.chb.setEnabled(False)
            else:
                self.chemical_section.hide()


    def btn_grant_code_search(self):
        # Gather the data from the form
        field_values = []

        for field in self.grant_code_search_panel.findChildren(FormLabelTextWidget):
            field_values.append(field.txt.text())

        # Search the database for the grant code by code
        grant_codes_by_name = get_grant_code_info_by_grant_code_name(field_values[0])

        # Search the database for the grant code by owner
        grant_codes_by_owner = get_grant_code_info_by_grant_code_owner(field_values[0])

        choice_selection_dialog = ChoiceSelectionDialogTwo("Multiple Grant Codes Found", "Multiple grant codes were found matching your search. Please select one:", "Grant Codes by Name:", "Grant Codes by Owner:", [grant_code[1] for grant_code in grant_codes_by_name], [grant_code[1] for grant_code in grant_codes_by_owner])
        if choice_selection_dialog.exec() == QDialog.Accepted:
            selected_optn = choice_selection_dialog.selected_option
            selected_idx = choice_selection_dialog.selected_idx

            if selected_optn == 1:
                selected_grant_code = grant_codes_by_name[selected_idx]
            else:
                selected_grant_code = grant_codes_by_owner[selected_idx]

            self.grant_code_id = selected_grant_code[0] 
            
            for idx, field in enumerate(self.grant_code_search_panel.findChildren(FormLabelTextWidget)):
                field.txt.setText(selected_grant_code[idx + 1])


    def open_add_grant_code_dialog(self):
        add_grant_code_dialog = AddGrantCodesDialog()
        add_grant_code_dialog.exec()


    def btn_proceed(self):
        # Gather all the information for the form
        item_info = get_item_info_by_id(self.item_id)
        supplier_info = get_supplier_info_by_id(item_info[2])
        grant_code_info = get_grant_code_info_by_id(self.grant_code_id)
        user_info = get_user_info_by_username(controller.logged_in_user)

        wb = load_workbook(resource_path("templates/Engineering_Order_Form_Template.xlsx"))
        ws = wb.active

        # Populate the supplier details
        ws[f"B10"] = supplier_info[1]
        ws[f"B11"] = supplier_info[3]
        ws[f"B12"] = supplier_info[4]
        ws[f"B13"] = supplier_info[5]
        ws[f"B14"] = supplier_info[6]
        ws[f"B16"] = supplier_info[7]
        ws[f"B17"] = supplier_info[8]
        ws[f"B18"] = supplier_info[9]
        ws[f"A50"] = supplier_info[2]
        ws[f"A53"] = "This is a trusted supplier."

        # Populate delivery details
        ws[f"F10"] = user_info[1]
        ws[f"F11"] = user_info[2]
        ws[f"F13"] = user_info[3]
        ws[f"F14"] = "Engineering Service Yard"
        ws[f"F15"] = "Swansea University Bay Campus"
        ws[f"F16"] = "Skewen"
        ws[f"F17"] = "Swansea"
        ws[f"F18"] = "SA1 8EN"

        # Populate the item details
        ws[f"B27"] = item_info[3]
        ws[f"C27"] = item_info[4]
        ws[f"H27"] = item_info[8]

        # Populate the grant code details
        ws[f"A56"] = grant_code_info[1]
        ws[f"H55"] = datetime.now().strftime("%d/%m/%Y")
        ws[f"H56"] = user_info[1]
        ws[f"H57"] = grant_code_info[2]

        wb.save(f"{db_info.SAVE_PATH}Engineering_Order_Form_{supplier_info[1]}.xlsx")


