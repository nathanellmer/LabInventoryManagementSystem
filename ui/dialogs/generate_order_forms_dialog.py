from openpyxl import load_workbook
from datetime import datetime
from PySide6.QtWidgets import QDialog, QHBoxLayout, QVBoxLayout, QListWidget
from PySide6.QtCore import Qt
from ui.custom_widgets.general.header_widgets import FormHeaderWidget
from ui.custom_widgets.general.footer_widgets import FormFooterWidget
from ui.custom_widgets.general.button_widgets import MainMenuButton, MainMenuLinkButton
from ui.custom_widgets.general.form_widgets import FormSearchPanelWidget, FormLabelTextWidget
from ui.dialogs.add_grant_codes_dialog import AddGrantCodesDialog
from ui.dialogs.choice_selection_dialog import ChoiceSelectionDialogTwo, ChoiceSelectionDialogFour
from ui.dialogs.show_item_dialog import ShowItemDialog
from core.control_functions import controller
from core.db_get_functions import get_item_info_by_product_code, get_item_info_by_supplier, get_item_info_by_description, get_item_info_by_name, get_items_info_by_reorder_flag, get_supplier_info_by_id, get_storage_location_info_by_id, get_grant_code_info_by_grant_code_name, get_grant_code_info_by_grant_code_owner, get_item_info_by_id, get_user_info_by_username, get_grant_code_info_by_id
from core.utility_functions import resource_path


class GenerateOrderFormsDialog(QDialog):
    def __init__(self):
        super().__init__()

        # Set the window size and title
        self.resize(500, 400)
        self.setWindowTitle("Generate Order Forms")

        # Initialise the GenerateOrderFormsDialog layout and its margins (left, top, right, bottom)
        gen_order_forms_dialog_layout = QVBoxLayout()
        gen_order_forms_dialog_layout.setContentsMargins(20, 20, 20, 20)

        # Add the header widget to the generate order forms dialog layout
        gen_order_forms_dialog_layout.addWidget(FormHeaderWidget("Generate Order Forms:"))

        # Add stretch
        gen_order_forms_dialog_layout.addStretch()

        # Add a search panel
        item_search_layout = QHBoxLayout()
        item_search_layout.addStretch()
        self.item_search_panel = FormSearchPanelWidget("Search for an item to add:", "Name or other information:", self.btn_item_search)
        item_search_layout.addWidget(self.item_search_panel)
        item_search_layout.addStretch()
        gen_order_forms_dialog_layout.addLayout(item_search_layout)

        # Add stretch
        gen_order_forms_dialog_layout.addStretch()

        # Add a list widget to display the items added to the order form
        self.list_items_id = []
        self.order_form_list = QListWidget()
        self.order_form_list.setSelectionMode(QListWidget.MultiSelection)
        self.order_form_list.doubleClicked.connect(self.open_show_item_dialog)
        self.load_reorder_items()
        gen_order_forms_dialog_layout.addWidget(self.order_form_list)

        # Add spacing
        gen_order_forms_dialog_layout.addSpacing(10)

        # Add buttons for removing selected items and generating the order form
        gen_order_forms_dialog_layout.addWidget(MainMenuButton("Remove Selected Items", self.btn_remove_selected_items), alignment=Qt.AlignCenter)

        # Add stretch
        gen_order_forms_dialog_layout.addStretch()

        # Add a search panel for the grant codes
        grant_code_search_layout = QHBoxLayout()
        grant_code_search_layout.addStretch()
        self.grant_code_search_panel = FormSearchPanelWidget("Search for a grant code:", "Grant code or owner:", self.btn_grant_code_search)
        grant_code_search_layout.addWidget(self.grant_code_search_panel)
        grant_code_search_layout.addStretch()
        gen_order_forms_dialog_layout.addLayout(grant_code_search_layout)

        # Add a link button widget for missing grant code information
        self.btn_missing_grant_code = MainMenuLinkButton("I cannot find the grant code...", self.open_add_grant_code_dialog)
        gen_order_forms_dialog_layout.addWidget(self.btn_missing_grant_code)

        # Add spacing
        gen_order_forms_dialog_layout.addSpacing(10)

        # Add button for order form generation
        gen_order_forms_dialog_layout.addWidget(MainMenuButton("Generate Order Form", self.btn_generate_order_form), alignment=Qt.AlignCenter)

        # Add stretch
        gen_order_forms_dialog_layout.addStretch()

        # Add the universal footer widget to the generate order forms dialog layout
        controller.close_all_windows.connect(self.close)
        gen_order_forms_dialog_layout.addWidget(FormFooterWidget())

        # Set the main layout for the window
        self.setLayout(gen_order_forms_dialog_layout)


    def load_reorder_items(self):
        # Get the list of items that need to be reordered
        reorder_items = get_items_info_by_reorder_flag()

        for item in reorder_items:
            supplier_name = get_supplier_info_by_id(item[2])[1]
            self.order_form_list.addItem(f"{item[1]} - {supplier_name} - {item[3]}")
            self.list_items_id.append(item[0])
        

    def btn_item_search(self):
        # Gather the data from the form
        field_values = []

        for field in self.item_search_panel.findChildren(FormLabelTextWidget):
            field_values.append(field.txt.text())

        # Search the database for the item by name
        items_by_name = get_item_info_by_name(field_values[0])

        # Search the database for the item by supplier
        items_by_supplier = get_item_info_by_supplier(field_values[0])

        # Search the database for the item by product code
        items_by_ref = get_item_info_by_product_code(field_values[0])

        # Search the database for the item by description
        items_by_description = get_item_info_by_description(field_values[0])

        choice_selection_dialog = ChoiceSelectionDialogFour("Multiple Items Found", "Multiple items were found matching your search. Please select one:", "Items by Name:", "Items by Supplier:", "Items by Product Code:", "Items by Description:", [item[1] for item in items_by_name], [item[1] for item in items_by_supplier], [item[1] for item in items_by_ref], [item[1] for item in items_by_description])
        if choice_selection_dialog.exec() == QDialog.Accepted:
            selected_optn = choice_selection_dialog.selected_option
            selected_idx = choice_selection_dialog.selected_idx

            if selected_optn == 1:
                selected_item = items_by_name[selected_idx]
            elif selected_optn == 2:
                selected_item = items_by_supplier[selected_idx]
            elif selected_optn == 3:
                selected_item = items_by_ref[selected_idx]
            elif selected_optn == 4:
                selected_item = items_by_description[selected_idx]

            # Add item to list widget
            supplier_name = get_supplier_info_by_id(selected_item[2])[1]
            self.order_form_list.addItem(f"{selected_item[1]} - {supplier_name} - {selected_item[3]}")
            self.list_items_id.append(selected_item[0])


    def open_show_item_dialog(self, index):
        selected_item_id = self.list_items_id[index.row()]
        
        show_item_dialog = ShowItemDialog(item_id=selected_item_id)
        show_item_dialog.exec()


    def btn_remove_selected_items(self):
        selected_items = self.order_form_list.selectedItems()
        for item in selected_items:
            idx = self.order_form_list.row(item)
            self.order_form_list.takeItem(idx)
            del self.list_items_id[idx]


    def btn_grant_code_search(self):
        # Gather the data from the form
        field_values = []

        for field in self.grant_code_search_panel.findChildren(FormLabelTextWidget):
            field_values.append(field.txt.text())

        # Search the database for the grant code by code
        grant_codes_by_name = get_grant_code_info_by_grant_code_name(field_values[0])

        # Search the database for the grant code by owner
        grant_codes_by_owner = get_grant_code_info_by_grant_code_owner(field_values[0])

        choice_selection_dialog = ChoiceSelectionDialogTwo("Multiple Grant Codes Found", "Multiple grant codes were found matching your search. Please select one:", "Grant Codes by Name:", "Grant Codes by Owner:", [grant_code[1] for grant_code in grant_codes_by_name], [grant_code[1] for grant_code in grant_codes_by_owner])
        if choice_selection_dialog.exec() == QDialog.Accepted:
            selected_optn = choice_selection_dialog.selected_option
            selected_idx = choice_selection_dialog.selected_idx

            if selected_optn == 1:
                selected_grant_code = grant_codes_by_name[selected_idx]
            else:
                selected_grant_code = grant_codes_by_owner[selected_idx]

            self.grant_code_id = selected_grant_code[0] 
            
            for idx, field in enumerate(self.grant_code_search_panel.findChildren(FormLabelTextWidget)):
                field.txt.setText(selected_grant_code[idx + 1])


    def open_add_grant_code_dialog(self):
        add_grant_code_dialog = AddGrantCodesDialog()
        add_grant_code_dialog.exec()


    def btn_generate_order_form(self):
        supplier_list = []
        for item_id in self.list_items_id:
            item_info = get_item_info_by_id(item_id)
            supplier_list.append(item_info[2])

        sorted_idx = sorted(range(len(supplier_list)), key=lambda k: supplier_list[k])
        sorted_item_ids = [self.list_items_id[i] for i in sorted_idx]
        sorted_supplier_list = [supplier_list[i] for i in sorted_idx]

        group = 0
        grouped_item_ids = []
        grouped_supplier_ids = []
        for i in range(len(sorted_supplier_list)):
            if i == 0:
                group += 1
                grouped_item_ids.append([sorted_item_ids[i]])
                grouped_supplier_ids.append(sorted_supplier_list[i])
            else:
                if sorted_supplier_list[i] == sorted_supplier_list[i-1]:
                    grouped_item_ids[group-1].append(sorted_item_ids[i])
                else:
                    group += 1
                    grouped_item_ids.append([sorted_item_ids[i]])
                    grouped_supplier_ids.append(sorted_supplier_list[i])

        print(grouped_supplier_ids)
        
        for idx, item_ids_group in enumerate(grouped_item_ids):
            self.populate_order_form(item_ids_group, grouped_supplier_ids[idx])


    def populate_order_form(self, item_ids, supplier_id):
        # Gather all the information for the form
        supplier_info = get_supplier_info_by_id(supplier_id)
        grant_code_info = get_grant_code_info_by_id(self.grant_code_id)
        user_info = get_user_info_by_username(controller.logged_in_user)

        wb = load_workbook(resource_path("templates/Engineering_Order_Form_Template.xlsx"))
        ws = wb.active

        # Populate the supplier details
        ws[f"B10"] = supplier_info[1]
        ws[f"B11"] = supplier_info[3]
        ws[f"B12"] = supplier_info[4]
        ws[f"B13"] = supplier_info[5]
        ws[f"B14"] = supplier_info[6]
        ws[f"B16"] = supplier_info[7]
        ws[f"B17"] = supplier_info[8]
        ws[f"B18"] = supplier_info[9]
        ws[f"A50"] = supplier_info[2]
        ws[f"A53"] = "This is a trusted supplier."

        # Populate delivery details
        ws[f"F10"] = user_info[1]
        ws[f"F11"] = user_info[2]
        ws[f"F13"] = user_info[3]
        ws[f"F14"] = "Engineering Service Yard"
        ws[f"F15"] = "Swansea University Bay Campus"
        ws[f"F16"] = "Skewen"
        ws[f"F17"] = "Swansea"
        ws[f"F18"] = "SA1 8EN"

        # Populate the grant code details
        ws[f"A56"] = grant_code_info[1]
        ws[f"H55"] = datetime.now().strftime("%d/%m/%Y")
        ws[f"H56"] = user_info[1]
        ws[f"H57"] = grant_code_info[2]

        for idx, item_id in enumerate(item_ids):
            item_info = get_item_info_by_id(item_id)
            
            # Populate the item details
            row = 27 + idx
            ws[f"B{row}"] = item_info[3]
            ws[f"C{row}"] = item_info[4]
            ws[f"H{row}"] = item_info[8]

        wb.save(f"Engineering_Order_Form_{supplier_info[1]}.xlsx")
